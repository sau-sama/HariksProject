


from django.db.models import  Count, Avg
from django.shortcuts import render, redirect
from django.db.models import Count
from django.db.models import Q
import datetime
import xlwt
from django.http import HttpResponse


import numpy as np
import pandas as pd
import seaborn as sns
sns.set(color_codes=True)

from sklearn.model_selection import train_test_split
from sklearn import metrics
from sklearn.metrics import accuracy_score,confusion_matrix
from sklearn.metrics import classification_report
from sklearn.feature_selection import mutual_info_classif

from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import RandomForestClassifier

from sklearn.preprocessing import StandardScaler
import openpyxl

# Create your views here.
from Remote_User.models import ClientRegister_Model,Loan_Status_Prediction,detection_ratio,detection_accuracy


def serviceproviderlogin(request):
    if request.method  == "POST":
        admin = request.POST.get('username')
        password = request.POST.get('password')
        if admin == "Admin" and password =="Admin":
            return redirect('View_Remote_Users')

    return render(request,'SProvider/serviceproviderlogin.html')


def viewtreandingquestions(request,chart_type):
    dd = {}
    pos,neu,neg =0,0,0
    poss=None
    topic = Loan_Status_Prediction.objects.values('ratings').annotate(dcount=Count('ratings')).order_by('-dcount')
    for t in topic:
        topics=t['ratings']
        pos_count=Loan_Status_Prediction.objects.filter(topics=topics).values('names').annotate(topiccount=Count('ratings'))
        poss=pos_count
        for pp in pos_count:
            senti= pp['names']
            if senti == 'positive':
                pos= pp['topiccount']
            elif senti == 'negative':
                neg = pp['topiccount']
            elif senti == 'nutral':
                neu = pp['topiccount']
        dd[topics]=[pos,neg,neu]
    return render(request,'SProvider/viewtreandingquestions.html',{'object':topic,'dd':dd,'chart_type':chart_type})

def View_All_Antifraud_Model_for_Internet_Loan_Prediction(request):

    obj = Loan_Status_Prediction.objects.all()
    return render(request, 'SProvider/View_All_Antifraud_Model_for_Internet_Loan_Prediction.html', {'objs': obj})

def Find_Internet_Loan_Prediction_Type_Ratio(request):
    detection_ratio.objects.all().delete()
    ratio = ""
    kword = '1'
    val = "Accepted"
    print(kword)
    obj = Loan_Status_Prediction.objects.all().filter(Q(prediction_rfc=kword))
    obj1 = Loan_Status_Prediction.objects.all()
    count = obj.count();
    count1 = obj1.count();
    ratio = (count / count1) * 100
    if ratio != 0:
        detection_ratio.objects.create(names=val, ratio=ratio)

    ratio1 = ""
    kword1 = '0'
    val1="Not Accepted"
    print(kword1)
    obj1 = Loan_Status_Prediction.objects.all().filter(Q(prediction_rfc=kword1))
    obj11 = Loan_Status_Prediction.objects.all()
    count1 = obj1.count();
    count11 = obj11.count();
    ratio1 = (count1 / count11) * 100
    if ratio1 != 0:
        detection_ratio.objects.create(names=val1, ratio=ratio1)


    obj = detection_ratio.objects.all()
    return render(request, 'SProvider/Find_Internet_Loan_Prediction_Type_Ratio.html', {'objs': obj})


def View_Remote_Users(request):
    obj=ClientRegister_Model.objects.all()
    return render(request,'SProvider/View_Remote_Users.html',{'objects':obj})

def ViewTrendings(request):
    topic = Loan_Status_Prediction.objects.values('topics').annotate(dcount=Count('topics')).order_by('-dcount')
    return  render(request,'SProvider/ViewTrendings.html',{'objects':topic})

def negativechart(request,chart_type):
    dd = {}
    pos, neu, neg = 0, 0, 0
    poss = None
    topic = Loan_Status_Prediction.objects.values('ratings').annotate(dcount=Count('ratings')).order_by('-dcount')
    for t in topic:
        topics = t['ratings']
        pos_count = Loan_Status_Prediction.objects.filter(topics=topics).values('names').annotate(topiccount=Count('ratings'))
        poss = pos_count
        for pp in pos_count:
            senti = pp['names']
            if senti == 'positive':
                pos = pp['topiccount']
            elif senti == 'negative':
                neg = pp['topiccount']
            elif senti == 'nutral':
                neu = pp['topiccount']
        dd[topics] = [pos, neg, neu]
    return render(request,'SProvider/negativechart.html',{'object':topic,'dd':dd,'chart_type':chart_type})

def charts(request,chart_type):
    chart1 = detection_ratio.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request,"SProvider/charts.html", {'form':chart1, 'chart_type':chart_type})

def charts1(request,chart_type):
    chart1 = detection_accuracy.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request,"SProvider/charts1.html", {'form':chart1, 'chart_type':chart_type})

def likeschart(request,like_chart):
    charts =detection_accuracy.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request,"SProvider/likeschart.html", {'form':charts, 'like_chart':like_chart})

def likeschart1(request,like_chart):
    charts =detection_ratio.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request,"SProvider/likeschart1.html", {'form':charts, 'like_chart':like_chart})

def Download_Trained_DataSets(request):

    response = HttpResponse(content_type='application/ms-excel')
    # decide file name
    response['Content-Disposition'] = 'attachment; filename="TrainedData.xls"'
    # creating workbook
    wb = xlwt.Workbook(encoding='utf-8')
    # adding sheet
    ws = wb.add_sheet("sheet1")
    # Sheet header, first row
    row_num = 0
    font_style = xlwt.XFStyle()
    # headers are bold
    font_style.font.bold = True
    # writer = csv.writer(response)
    obj = Loan_Status_Prediction.objects.all()
    data = obj  # dummy method to fetch data.
    for my_row in data:
        row_num = row_num + 1
        ws.write(row_num, 0, my_row.Idno, font_style)
        ws.write(row_num, 1, my_row.Age, font_style)
        ws.write(row_num, 2, my_row.Experience, font_style)
        ws.write(row_num, 3, my_row.Income, font_style)
        ws.write(row_num, 4, my_row.ZIP_Code, font_style)
        ws.write(row_num, 5, my_row.Family, font_style)
        ws.write(row_num, 6, my_row.CCAvg, font_style)
        ws.write(row_num, 7, my_row.Education, font_style)
        ws.write(row_num, 8, my_row.Mortgage, font_style)
        ws.write(row_num, 9, my_row.Securities_Account, font_style)
        ws.write(row_num, 10, my_row.CD_Account, font_style)
        ws.write(row_num, 11, my_row.Online, font_style)
        ws.write(row_num, 12, my_row.CreditCard, font_style)
        ws.write(row_num, 13, my_row.prediction_svm, font_style)
        ws.write(row_num, 14, my_row.prediction_Logistic, font_style)
        ws.write(row_num, 15, my_row.prediction_rfc, font_style)
        ws.write(row_num, 16, my_row.prediction_dtc, font_style)
        ws.write(row_num, 17, my_row.prediction_knc, font_style)

    wb.save(response)
    return response

def Train_Test_DataSets(request):

    detection_accuracy.objects.all().delete()

    Bank = pd.read_csv("Bank_Datasets.csv")
    Target = ["Personal_Loan"]
    t = Bank[Target]
    t.head()
    loan_acceptance_count = pd.DataFrame(Bank['Personal_Loan'].value_counts()).reset_index()
    loan_acceptance_count.columns = ['Labels', 'Personal_Loan']
    loan_acceptance_count
    pie_labels = loan_acceptance_count['Labels']
    pie_labels = ['Not Accepted' if x == 0 else 'Accepted' for x in pie_labels]
    pie_data = loan_acceptance_count['Personal_Loan']
    explode = (0, 0.15)
    wp = {'linewidth': 1, 'edgecolor': '#000000'}

    def func(pct, allvalues):
        absolute = int(np.round(pct / 100. * np.sum(allvalues)))
        return "{:.1f}%\n({:d})".format(pct, absolute)

    categorical_variables = [col for col in Bank.columns if Bank[col].nunique() <= 5]
    print(categorical_variables)
    categorical_variables.remove("Personal_Loan")
    print(categorical_variables)

    X = Bank.drop('Personal_Loan', axis=1)  # set X with all feature except Personal Loan
    Y = Bank[['Personal_Loan']]  # set y with our target feature Personal Loan
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.20, random_state=1, stratify=Y)


    print("SVM")
    # SVM Model
    from sklearn import svm

    lin_clf = svm.LinearSVC()
    lin_clf.fit(X_train, y_train)
    predict_svm = lin_clf.predict(X_test)
    svm_acc = accuracy_score(y_test, predict_svm) * 100
    print(svm_acc)
    from sklearn.metrics import confusion_matrix, f1_score
    print(confusion_matrix(y_test, predict_svm))
    print(classification_report(y_test, predict_svm))
    detection_accuracy.objects.create(names="SVM", ratio=svm_acc)
    # Logistic Regression Model
    print("Logistic Regression")
    from sklearn.linear_model import LogisticRegression
    logreg = LogisticRegression(random_state=42)
    logreg.fit(X_train, y_train)
    predict_log = logreg.predict(X_test)
    logistic = accuracy_score(y_test, predict_log) * 100
    print(logistic)
    from sklearn.metrics import confusion_matrix, f1_score
    print(confusion_matrix(y_test, predict_log))
    print(classification_report(y_test, predict_log))
    detection_accuracy.objects.create(names="Logistic Regression", ratio=logistic)

    # Decision Tree Classifier
    print("Decision Tree Classifier")
    dtc = DecisionTreeClassifier()
    dtc.fit(X_train, y_train)
    dtcpredict = dtc.predict(X_test)
    print("ACCURACY")
    print(accuracy_score(y_test, dtcpredict) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(y_test, dtcpredict))
    print("CONFUSION MATRIX")
    print(confusion_matrix(y_test, dtcpredict))
    detection_accuracy.objects.create(names="Decision Tree Classifier", ratio=accuracy_score(y_test, dtcpredict) * 100)

    # Random Forest Classifier
    print("Random Forest Classifier")
    from sklearn.ensemble import RandomForestClassifier
    RFC = RandomForestClassifier(random_state=0)
    RFC.fit(X_train, y_train)
    pred_rfc = RFC.predict(X_test)
    RFC.score(X_test, y_test)
    print("ACCURACY")
    print(accuracy_score(y_test, pred_rfc) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(y_test, pred_rfc))
    print("CONFUSION MATRIX")
    print(confusion_matrix(y_test, pred_rfc))
    detection_accuracy.objects.create(names="Random Forest Classifier", ratio=accuracy_score(y_test, pred_rfc) * 100)

    print("KNeighborsClassifier")
    from sklearn.neighbors import KNeighborsClassifier
    kn = KNeighborsClassifier()
    kn.fit(X_train, y_train)
    knpredict = kn.predict(X_test)
    print("ACCURACY")
    print(accuracy_score(y_test, knpredict) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(y_test, knpredict))
    print("CONFUSION MATRIX")
    print(confusion_matrix(y_test, knpredict))
    detection_accuracy.objects.create(names="KNeighbors Classifier", ratio=accuracy_score(y_test, knpredict) * 100)


    prediction_df = pd.DataFrame(columns=[])
    prediction_df['Idno'] = X_test.Idno
    prediction_df['Age'] = X_test.Age
    prediction_df['Experience'] = X_test.Experience
    prediction_df['Income'] = X_test.Income
    prediction_df['ZIP_Code'] = X_test.ZIP_Code
    prediction_df['Family'] = X_test.Family
    prediction_df['CCAvg'] = X_test.CCAvg
    prediction_df['Education'] = X_test.Education
    prediction_df['Mortgage'] = X_test.Mortgage
    prediction_df['Securities_Account'] = X_test.Securities_Account
    prediction_df['CD_Account'] = X_test.CD_Account
    prediction_df['Online'] = X_test.Online
    prediction_df['CreditCard'] = X_test.CreditCard
    prediction_df['SVM'] = lin_clf.predict(X_test)
    prediction_df['Logistic Regression'] = logreg.predict(X_test)
    prediction_df['RandomForestClassifier'] = RFC.predict(X_test)
    prediction_df['Decision Tree Classifier'] = dtc.predict(X_test)
    prediction_df['KNeighborsClassifier'] = kn.predict(X_test)

    prediction_df.to_excel('Loan_approval_prediction.xlsx', index=False)

    excel_file = ("Loan_approval_prediction.xlsx")
    # you may put validations here to check extension or file size
    wb = openpyxl.load_workbook(excel_file)
    # getting all sheets
    sheets = wb.sheetnames
    print(sheets)
    # getting a particular sheet
    worksheet = wb["Sheet1"]
    print(worksheet)
    # getting active sheet
    active_sheet = wb.active
    print(active_sheet)
    # reading a cell
    print(worksheet["A1"].value)
    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
            print(cell.value)
        excel_data.append(row_data)
        Loan_Status_Prediction.objects.all().delete()
    for r in range(1, active_sheet.max_row + 1):
        Loan_Status_Prediction.objects.create(
        Idno=active_sheet.cell(r, 1).value,
        Age=active_sheet.cell(r, 2).value,
        Experience=active_sheet.cell(r, 3).value,
        Income=active_sheet.cell(r, 4).value,
        ZIP_Code=active_sheet.cell(r, 5).value,
        Family=active_sheet.cell(r, 6).value,
        CCAvg=active_sheet.cell(r, 7).value,
        Education=active_sheet.cell(r, 8).value,
        Mortgage=active_sheet.cell(r, 9).value,
        Securities_Account=active_sheet.cell(r, 10).value,
        CD_Account=active_sheet.cell(r, 11).value,
        Online=active_sheet.cell(r, 12).value,
        CreditCard=active_sheet.cell(r, 13).value,
        prediction_svm=active_sheet.cell(r, 14).value,
        prediction_Logistic=active_sheet.cell(r, 15).value,
        prediction_rfc=active_sheet.cell(r, 16).value,
        prediction_dtc=active_sheet.cell(r, 17).value,
        prediction_knc=active_sheet.cell(r, 18).value,

        )

    obj = detection_accuracy.objects.all()

    return render(request,'SProvider/Train_Test_DataSets.html', {'objs': obj,'objs1': excel_data})














